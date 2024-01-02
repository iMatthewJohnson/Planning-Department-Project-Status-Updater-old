
import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def update_comments(project_status_workbook, status_responses_workbook, column_letters):
    ps_wb = load_workbook(filename=project_status_workbook)
    sr_wb = load_workbook(filename=status_responses_workbook)

    sr_ws = sr_wb.active

    # Helper function to get letter value from dictionary and convert into the column number
    get_index = lambda col_letter: column_index_from_string(column_letters[col_letter]) - 1

    # Mapping of status letters to the corresponding columns in the project status workbook
    status_to_col = {
        'A': get_index('status_a'),
        'B': get_index('status_b'),
        'C': get_index('status_c'),
        'D': get_index('status_d'),
        'E': get_index('status_e'),
        'F': get_index('status_f')
    }

    # Iterate over action ID and comments in the responses workbook
    for sr_row in sr_ws.iter_rows(min_row=2, values_only=True):
        action_id = sr_row[get_index('action_ID_response')].lower()
        comment = sr_row[get_index('comments_response')]
        status_letter = sr_row[get_index('status_response')][0] if sr_row[get_index('status_response')] else ''  # Get first char of status, check if not None
        found = False

        # Go through each sheet in the project status workbook
        if not found:
            for sheet in ps_wb.sheetnames:
                if sheet in ["Merge", "Full Plan", "OWNERS LIST"]:  # Skip these sheets
                    continue
                ps_ws = ps_wb[sheet]
                for ps_row in ps_ws.iter_rows(min_row=2):
                    # Look for the matching action ID
                    if ps_row[get_index('action_ID_project')].value and ps_row[get_index('action_ID_project')].value.lower() == action_id:
                        # Clear existing statuses
                        for col_index in status_to_col.values():
                            ps_row[col_index].value = None

                        # Set new status
                        if status_letter in status_to_col:
                            new_status_col_index = status_to_col[status_letter]
                            ps_row[new_status_col_index].value = 'X'
                            found = True
                            break
                if found:
                    ps_row[get_index('comments_project')].value = comment
                    break

    ps_wb.save(filename=project_status_workbook)


def duplicate_workbook(file_path):
    file_dir, file_name = os.path.split(file_path)
    name, ext = os.path.splitext(file_name)
    old_file_path = os.path.join(file_dir, f"{name} (old){ext}")
    shutil.move(file_path, old_file_path)
    shutil.copy2(old_file_path, file_path) # Use copy2 to preserve metadata


# Main function to run the program
def run_sync(project_status_wb_path, status_responses_wb_path, column_letters):
    duplicate_workbook(project_status_wb_path)
    update_comments(project_status_wb_path, status_responses_wb_path, column_letters)